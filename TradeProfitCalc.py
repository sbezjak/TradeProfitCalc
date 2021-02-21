#!Python 3.8.0

# Guide:
# 1. Download a CSV report from Trading212 for a desired time frame. History -> Export History 
# 2. Enter file path name.
# 3. Optionaly rename the file that will be created with this.
# 4. Build (Ctrl+B)
# 5. $ See your profit $

from openpyxl import load_workbook
import pandas as pd

# Convert from csv to xlsx
# --Enter export file path name here--
pd.read_csv('C:/Users/Sara/Desktop/from_2021-01-25_to_2021-01-29_MTYxMjAwMzMzNjkwNw.csv').to_excel('C:/Users/Sara/Desktop/excel.xlsx')
workbook = load_workbook(filename="excel.xlsx")
workbook.sheetnames
# Open sheet
sheet = workbook.active
# Create a dictionary of column names
ColNames = {}
Current  = 0
for column in sheet.iter_cols(1, sheet.max_column):
    ColNames[column[0].value] = Current
    Current += 1
# Now you can access data by column name
resultSum = 0
i = 0
for row_cells in sheet.iter_rows(min_row=1):
	i = i + 1
	if "sell" in row_cells[ColNames['Action']].value:
		resultSum = resultSum + row_cells[ColNames['Result (EUR)']].value
print(resultSum)

# Write result sum to a new cell in new excel file
sheet['K' + str(i + 1)] = resultSum
sheet.cell(row = i + 1, column = 11)
# --Enter output file name here--
workbook.save('Trades_2021-02-16_to_2021-02-19.xlsx')